import tkinter as tk
import openpyxl
import HomeView as Home
import xlrd
import os

LARGE_FONT= ("Verdana", 15)
DaTa = "comming soon "


class DeleteItemsView(tk.Frame):
    
    def delete_item_fun(self,bottomframe,imagedestination,controller):    
        os.remove("image/"+imagedestination)
        filename = imagedestination.split('/')
        xfile = openpyxl.load_workbook("db/"+filename[0]+".xlsx")
        ws = xfile.active
        deleted = 0 
        for xx in range(ws.max_row):
            if ws.cell(row=(xx+1),column=1).value == imagedestination:
                #print("delete ",Lb1.get(Lb1.curselection())
                ws.cell(row=(xx+1),column=1).value = '' 
                ws.cell(row=(xx+1),column=2).value = '' 
                deleted= 1
                
        Message_label = ""
        if deleted == 1:
            Message_label = "Deletd field "+imagedestination
        else:
                Message_label="something wrong!"
        label = tk.Label(bottomframe, text=Message_label)
        label.pack(pady=2,padx=2)
        xfile.save("db/"+filename[0]+".xlsx")
        controller.Delete_items(self.wow)
        
        
    #render items button
    def create_button(self,bottomframe,imagedestination,controller):
        photo = tk.PhotoImage(file = r"image/"+imagedestination)
        label = tk.Label(image=photo)
        label.image = photo
        bb=tk.Button(bottomframe, 
                   image = photo,text='Convert to voice',
                  command=lambda:self.delete_item_fun(bottomframe,imagedestination,controller))
        return bb

    
    #render items from db
    def __init__(self, parent,wow, controller):
        
        self.wow = wow

        wb = xlrd.open_workbook("db/listofdatatable.xlsx")
        sheet = wb.sheet_by_index(0)

        excelname = sheet.cell_value(wow,0)
        
        wb = xlrd.open_workbook("db/"+excelname+".xlsx")
        sheet = wb.sheet_by_index(0)
        Emotion ={}
        
        for x in range(sheet.nrows):
            if sheet.cell_value(x,0)!='':
                Emotion[sheet.cell_value(x,0)]=sheet.cell_value(x, 1)
            
        self.cc = controller
        tk.Frame.__init__(self, parent)
        label = tk.Label(self, text="DELETE ITEMS PAGE \n"+excelname.upper(), font=LARGE_FONT)
        label.pack(pady=10,padx=10)
        
        photo = tk.PhotoImage(file = r"image/back.png")
        label = tk.Label(image=photo)
        label.image = photo
        button1 = tk.Button(self, text="Back to Home",image = photo,
                            command=lambda: controller.show_frame(Home.Home))
        button1.pack()
        
        i =1
        bottomframe = tk.Frame(self)
        bottomframe.pack()
        for x in Emotion:
            
            b = self.create_button(bottomframe,x,controller)
            if i > 5:
                bottomframe = tk.Frame(self)
                bottomframe.pack()
                b.pack(side=tk.LEFT)
                i = 0
            else:
                b.pack(side=tk.LEFT)
                
            i=i+1
        
