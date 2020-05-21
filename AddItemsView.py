import HomeView as Home
import UploadImageController
import tkinter as tk
import random
import string
import openpyxl
from tkinter import filedialog
import os

LARGE_FONT= ("Verdana", 15)
DaTa = "comming soon "

class AddItems(tk.Frame):
    
    #create new field and item
    def __init__(self, parent,selectedField, controller):
        tk.Frame.__init__(self,parent)
        bottomframe = tk.Frame(self)
        bottomframe.pack()
        label = tk.Label(bottomframe, text="ADD ITEMS PAGE", font=LARGE_FONT)
        label.pack(pady=10,padx=10)

        photo = tk.PhotoImage(file = r"image/back.png")
        label = tk.Label(image=photo)
        label.image = photo
        button1 = tk.Button(bottomframe, text="Back to Home",image = photo,
                            command=lambda: controller.show_frame(Home.Home))
        button1.pack()

        label = tk.Label(bottomframe, text=selectedField.upper())
        label.pack(pady=5,padx=5)

        label = tk.Label(bottomframe, text="Upload Image")
        label.pack(pady=5,padx=5)
        
        photo = tk.PhotoImage(file = r"image/blackimage.png")
        label = tk.Label(image=photo)
        label.image = photo
        
        self.imagedestination = selectedField+"/"+self.randomString(10)+".png"
        button1 = tk.Button(bottomframe, text="Upload New Image",image = photo,
                            command=lambda: self.Upload_Image(self.imagedestination,bottomframe))
        button1.pack(pady=5,padx=5)

        label = tk.Label(bottomframe, text="Enter Speech")
        label.pack(pady=5,padx=5)
        EnterData = tk.Entry(bottomframe)
        EnterData.pack(pady=5,padx=5)

        button1 = tk.Button(bottomframe, text="Add Data",
                            command=lambda: self.Add_Data_button(bottomframe,EnterData,self.imagedestination))
        button1.pack(pady=5,padx=5)
        
    #submit field and item data to db
    def Add_Data_button(self,bottomframe,EnterData,imagedestination):
        if EnterData.get() != "":
            if os.path.exists("image/"+imagedestination):
                filename = imagedestination.split('/')
                xfile = openpyxl.load_workbook("db/"+filename[0]+".xlsx")
                ws = xfile.active
                
                check = 0
                for xx in range(ws.max_row):
                    if ws.cell(row=(xx+1),column=1).value == imagedestination:
                        check = 1
                        print("duplicate !! ")
                if check == 0:
                    
                    if  ws.cell(row=(ws.max_row),column=1).value==None:
                        MM = ws.max_row
                        ws.cell(row=(MM),column=1).value = imagedestination
                        ws.cell(row=(MM),column=2).value = EnterData.get()
                    else:
                        MM = ws.max_row+1
                        ws.cell(row=(MM),column=1).value = imagedestination
                        ws.cell(row=(MM),column=2).value = EnterData.get()
                    xfile.save("db/"+filename[0]+".xlsx")
                    label = tk.Label(bottomframe, text="Data Updated ! ")
                    label.pack(pady=2,padx=2)
                    self.imagedestination = filename[0]+"/"+self.randomString(10)+".png"
                    EnterData.insert(0, "")
            else:
                label = tk.Label(bottomframe, text="Please Upload Image !")
                label.pack(pady=2,padx=2)

        else:
            label = tk.Label(bottomframe, text="Please Enter data !")
            label.pack(pady=2,padx=2)
            
    #upload emoji for new field
    def Upload_Image(self,imagedestination,bottomframe):
        filename=filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("png files","*.png"),("jpeg files","*.jpg"),("all files","*.*")))
        if filename:
            UploadImageController.UploadImageController(filename,imagedestination)
            label = tk.Label(bottomframe, text="Image is updated! ")
            label.pack(pady=2,padx=2)
        else:
            print("no image")
            label = tk.Label(bottomframe, text="You have not select Image")
            label.pack(pady=2,padx=2)
    
    #create unique string for store emoji
    def randomString(self,stringLength=10):
        """Generate a random string of fixed length """
        letters = string.ascii_lowercase
        return ''.join(random.choice(letters) for i in range(stringLength))
