import HomeView as Home
import tkinter as tk
import xlrd
import AddFieldController
import os
import openpyxl
import glob

LARGE_FONT= ("Verdana", 15)
DaTa = "comming soon "

class AddFields(tk.Frame):
    
    #get field and according to that create items
    def __init__(self, parent,imagedestination, controller):
        tk.Frame.__init__(self,parent)
        bottomframe = tk.Frame(self)
        bottomframe.pack()
        label = tk.Label(bottomframe, text="ADD/DELETE PAGE", font=LARGE_FONT)
        label.pack(pady=10,padx=10)

        photo = tk.PhotoImage(file = r"image/back.png")
        label = tk.Label(image=photo)
        label.image = photo
        button1 = tk.Button(bottomframe, text="Back to Home",image = photo,
                            command=lambda: controller.show_frame(Home.Home))
        button1.pack()

        label = tk.Label(bottomframe, text="Create new field ")
        label.pack(pady=10,padx=10)
        
        EnterData = tk.Entry(bottomframe)
        EnterData.pack(pady=5,padx=5)

        button1 = tk.Button(bottomframe, text="CREATE FIELD",
                            command=lambda: self.create_excel_Database(EnterData,controller))
        button1.pack(pady=5,padx=5)

        wb = xlrd.open_workbook("db/listofdatatable.xlsx")
        sheet = wb.sheet_by_index(0)
        
        
        Lb1 = tk.Listbox(bottomframe)
        for i in range(sheet.nrows):
            if sheet.cell_value(i, 0)!='':
                Lb1.insert(i, sheet.cell_value(i, 0))

        
        Lb1.pack()

        button1 = tk.Button(bottomframe, text="CREATE ITEMS",
                            command=lambda: self.select_data(Lb1,controller))
        button1.pack(pady=5,padx=5)
        
        button1 = tk.Button(bottomframe, text="DELETE FIELD",
                            command=lambda: self.delete_field(bottomframe,Lb1,controller))
        button1.pack(pady=5,padx=5)
        
        button1 = tk.Button(bottomframe, text="DELETE ITEMS",
                            command=lambda: self.delete_item(Lb1,controller))
        button1.pack(pady=5,padx=5)
        
    #create new entry for fields into db
    def create_excel_Database(self,EnterData,controller):
        if EnterData.get()!= '':
            
            AddFieldController.AdFieldController(EnterData.get())

            controller.Add_items(EnterData.get())
            try:
                os.mkdir("image/"+EnterData.get())
            except:
                print("Fail already exist")

    #select field and create new item in that field
    def select_data(self,Lb1,controller):
        if Lb1.curselection():
            controller.Add_items(Lb1.get(Lb1.curselection()))
        else:
            print(" No selected  ")

    def delete_field(self,bottomframe,Lb1,controller):
        if Lb1.curselection():
            xfile = openpyxl.load_workbook("db/listofdatatable.xlsx")
            ws = xfile.active
            deleted = 0 
            for xx in range(ws.max_row):
                if ws.cell(row=(xx+1),column=1).value == Lb1.get(Lb1.curselection()):
                    #print("delete ",Lb1.get(Lb1.curselection())
                    ws.cell(row=(xx+1),column=1).value = '' 
                    deleted= 1
                    
            Message_label = ""
            if deleted == 1:
                
                files = glob.glob('image/'+Lb1.get(Lb1.curselection())+'/*.*', recursive=True)
                for f in files:
                    print("delete ",f)
                    os.remove(f)
                os.rmdir("image/"+Lb1.get(Lb1.curselection()))
                os.remove("db/"+Lb1.get(Lb1.curselection())+".xlsx")
                Message_label = "Deletd field "+Lb1.get(Lb1.curselection())
            else:
                Message_label="something wrong!"
            label = tk.Label(bottomframe, text=Message_label)
            label.pack(pady=2,padx=2)
            xfile.save("db/listofdatatable.xlsx")
            controller.Add_page_create()
            
        else:
            print(" No selected  ")
            
    
    def delete_item(self,Lb1,controller):
        if Lb1.curselection():
            wb = xlrd.open_workbook("db/listofdatatable.xlsx")
            sheet = wb.sheet_by_index(0)
            total_row = sheet.nrows
            
            for i in range(total_row):
                if sheet.cell_value(i, 0)==Lb1.get(Lb1.curselection()):
                    controller.Delete_items(i)
                    break;
        else:
            print(" No selected  ")
            
