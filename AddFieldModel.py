
import openpyxl

class AddFieldModel:
    
    #store new field into db
    def __init__(self,EnterData):
        xfile = openpyxl.load_workbook("db/listofdatatable.xlsx")
        ws = xfile.active
        check = 0
        for xx in range(ws.max_row):
            if ws.cell(row=(xx+1),column=1).value == EnterData:
                check = 1
        if check == 0:
            ws.cell(row=(ws.max_row+1),column=1).value = EnterData
            xfile.save("db/listofdatatable.xlsx")
            wb = openpyxl.Workbook()
            ws =  wb.active
            ws.title = "Changed Sheet"
            wb.save(filename = 'db/'+EnterData+'.xlsx')


        
