import tkinter as tk
import xlrd
import HomeView as Home
import openpyxl
from twilio.rest import Client

LARGE_FONT= ("Verdana", 15)
DaTa = "comming soon "


class EmergencyView(tk.Frame):
    
    #render edit page when click on edit items
    def __init__(self, parent, controller):
        tk.Frame.__init__(self,parent)
        bottomframe = tk.Frame(self)
        bottomframe.pack()
        label = tk.Label(bottomframe, text="EMERGENCY PAGE", font=LARGE_FONT)
        label.pack(pady=10,padx=10)

        photo = tk.PhotoImage(file = r"image/back.png")
        label = tk.Label(image=photo)
        label.image = photo
        button1 = tk.Button(bottomframe, text="Back to Home",image = photo,
                            command=lambda: controller.show_frame(Home.Home))
        button1.pack(pady=10,padx=10)
        
        wb = xlrd.open_workbook("db/emergencycontact.xlsx")
        sheet = wb.sheet_by_index(0)
        PhoneNumber= ""
        Message = ""
        for x in range(sheet.nrows):
            PhoneNumber =  sheet.cell_value(x,0)
            Message= sheet.cell_value(x,1)
            
        
        label = tk.Label(bottomframe, text="EMERGENCY NUMBER")
        label.pack(pady=10,padx=10)
        
        EnterData = tk.Entry(bottomframe)
        EnterData.insert(0, PhoneNumber)
        EnterData.pack(pady=10,padx=10)
        
        label = tk.Label(bottomframe, text="MESSAGE")
        label.pack(pady=5,padx=5)
        
        EnterData2 = tk.Entry(bottomframe,width=30)
        EnterData2.insert(0, Message)
        EnterData2.pack(pady=10,padx=10,ipady=10)
        
        button = tk.Button(bottomframe, text="SEND MESSAGE",
                            command=lambda:self.Send_Message(EnterData,EnterData2,bottomframe))    
        button.config( height = 0, width = 16 )
        button.pack(pady=10,padx=10)
       
   
    def Send_Message(self,EnterData,EnterData2,bottomframe):
        if EnterData.get()!='':
            if str.isnumeric(EnterData.get()) and len(EnterData.get())==10:
                print("valid phone ",EnterData.get())
                Message = EnterData2.get()
                if Message=='':
                    Message = "EMERGENCY !!!"
                print("message ",Message)
                xfile = openpyxl.load_workbook("db/emergencycontact.xlsx")
                ws = xfile.active
                ws.cell(row=(1),column=1).value= EnterData.get()
                ws.cell(row=(1),column=2).value= Message
                xfile.save("db/emergencycontact.xlsx")
                #client = Client("", "")
                #client.messages.create(to="+1"+EnterData.get(), from_="", body=Message)

                label = tk.Label(bottomframe, text="SMS send to Emergency Number! *Not Work Need to add account*")
                label.pack(pady=2,padx=2)
                
            else:
                label = tk.Label(bottomframe, text="Invalid Emergency Number")
                label.pack(pady=2,padx=2)
